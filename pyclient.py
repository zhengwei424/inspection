#!/bin/env python
# _*_ coding: utf-8 _*_

import os
from kubernetes import client, config
import tokens
from urllib3.exceptions import InsecureRequestWarning
from urllib3 import disable_warnings
from openpyxl import Workbook


def auth(mode, kubeconfig_file=None, **kwargs):
    if mode == 'kubeconfig':
        if os.path.isfile(kubeconfig_file):
            config.load_kube_config(config_file=kubeconfig_file)
    elif mode == 'token':
        if isinstance(kwargs.get('kwargs').get('token'), str):
            configuration = client.Configuration()
            configuration.host = kwargs.get('kwargs').get('host')
            configuration.api_key = {'authorization': 'Bearer ' + kwargs.get('kwargs').get('token')}
            configuration.verify_ssl = False
            client.Configuration.set_default(configuration)
            # 禁用安全请求警告
            disable_warnings(InsecureRequestWarning)
    else:
        print("Please set the authentication mode: kubeconfig or token")


if __name__ == '__main__':
    # 初始化变量
    cluster_name = ''
    node_ip = ''
    project_name = ''
    namespace = ''
    app_name = ''
    services_name = ''
    services_ports = ''
    rs_type = ''
    rs_name = ''
    pod_name = ''
    volumes = ''
    container_name = ''
    container_ports = ''
    container_resources = ''
    container_liveness_probe = ''
    container_readiness_probe = ''
    container_volume_mounts = ''

    # 初始化一个Workbook实例
    wb = Workbook()

    for cluster_index in range(len(tokens.production_envs)):
        if cluster_index != 0:
            wb.active = wb.create_sheet()
        ws = wb.active
        ws.title = str(tokens.production_envs[cluster_index].get("name"))
        ws['A1'] = 'cluster_name'
        ws['B1'] = 'host_ip'
        ws['C1'] = 'project_name'
        ws['D1'] = 'namespace'
        ws['E1'] = 'app_name'
        ws['F1'] = 'service_name'
        ws['G1'] = 'service_ports'
        ws['H1'] = 'replicas_type'
        ws['I1'] = 'replicas_name'
        ws['J1'] = 'pod_name'
        ws['K1'] = 'volumes'
        ws['L1'] = 'container'
        ws['M1'] = 'container_ports'
        ws['N1'] = 'resources'
        ws['O1'] = 'liveness_probe'
        ws['P1'] = 'readiness_probe'
        ws['Q1'] = 'volume_mounts'

        index = 2

        auth(mode='token', kwargs=tokens.production_envs[cluster_index])
        # 创建v1对象
        v1 = client.CoreV1Api()

        # 创建extensions/v1对象
        rs2 = client.ExtensionsV1beta1Api()

        # 集群名称
        cluster_name = tokens.production_envs[cluster_index].get("name")

        ns = v1.list_namespace().items
        for ns_index in range(len(ns)):
            # 命名空间
            namespace = ns[ns_index].metadata.name
            # 获取所有pod信息——>pods类型是list——>pods的每项参数类型是kubernetes.client.models.v1_pod.V1Pod(所有字段当做属性值通过点直接获取）
            pods = v1.list_namespaced_pod(namespace).items

            # 获取deployment名称信息列表
            deployments = rs2.list_namespaced_deployment(namespace).items
            deploy_names = [deployments[deploy_index].metadata.name for deploy_index in range(len(deployments))]

            # 获取services
            services = v1.list_namespaced_service(namespace).items

            # 获取service特定信息  名称:端口  字典
            services_info = {}
            for svc_index in range(len(services)):
                services_info[services[svc_index].metadata.name] = str(services[svc_index].spec.ports)

            # 获取endpoints
            endpoints = v1.list_namespaced_endpoints(namespace).items

            # 获取endpoints特定信息， pod_ip:endpoint_name（同service名） 字典
            ep_info = {}
            for ep_index in range(len(endpoints)):
                if endpoints[ep_index].subsets is not None:
                    addresses = endpoints[ep_index].subsets[0].addresses
                    for addr_index in range(len(addresses)):
                        ep_info[addresses[addr_index].ip] = endpoints[ep_index].metadata.name

            for pod_index in range(len(pods)):

                # 节点名称
                node_ip = pods[pod_index].status.host_ip
                # 项目名称
                if pods[pod_index].metadata.namespace in tokens.production_envs[cluster_index].get('projects').keys():
                    project_name = \
                        [i for i in
                         tokens.production_envs[cluster_index].get('projects').get(
                             pods[pod_index].metadata.namespace).keys()][0]
                else:
                    project_name = ''

                # 副本集类型和名称
                if pods[pod_index].metadata.owner_references is not None:
                    # 副本集类型
                    rs_type = pods[pod_index].metadata.owner_references[0].kind

                    # 副本集名称
                    rs_name = pods[pod_index].metadata.owner_references[0].name.split('-')
                    rs_name.pop(-1)
                    rs_name = '-'.join(rs_name)

                    if rs_type == 'ReplicaSet' and rs_name in deploy_names:
                        rs_type = "Deployment"
                        if tokens.production_envs[cluster_index].get('projects').get(namespace) is not None:
                            app_name = tokens.production_envs[cluster_index].get('projects').get(namespace).get(
                                project_name).get(rs_name)
                        else:
                            app_name = ''
                else:
                    rs_type = ''
                    rs_name = ''

                # service
                if pods[pod_index].status.pod_ip in ep_info.keys():
                    services_name = ep_info[pods[pod_index].status.pod_ip]
                    services_ports = str(services_info[services_name])
                else:
                    services_name = ''
                    services_ports = ''

                # pod名称
                pod_name = pods[pod_index].metadata.name

                # volumes
                volumes_test = pods[pod_index].spec.volumes
                volumes_list = []
                for volume_index in range(len(volumes_test)):
                    dic = eval(str(volumes_test[volume_index]))
                    for key in list(dic.keys()):
                        if dic[key] is None:
                            del dic[key]
                    volumes_list.append(dic)
                volumes = str(volumes_list)

                # 获取containers列表
                containers = pods[pod_index].spec.containers
                for container_index in range(len(containers)):
                    container_name = containers[container_index].name
                    container_ports = str(containers[container_index].ports)
                    container_resources = str(containers[container_index].resources)
                    container_liveness_probe = str(containers[container_index].liveness_probe)
                    container_readiness_probe = str(containers[container_index].readiness_probe)
                    container_volume_mounts = str(containers[container_index].volume_mounts)

                    ws['A' + str(index)] = cluster_name
                    ws['B' + str(index)] = node_ip
                    ws['C' + str(index)] = project_name
                    ws['D' + str(index)] = namespace
                    ws['E' + str(index)] = app_name
                    ws['F' + str(index)] = services_name
                    ws['G' + str(index)] = str(services_ports)
                    ws['H' + str(index)] = rs_type
                    ws['I' + str(index)] = rs_name
                    ws['J' + str(index)] = pod_name
                    ws['K' + str(index)] = str(volumes)
                    ws['L' + str(index)] = container_name
                    ws['M' + str(index)] = str(container_ports)
                    ws['N' + str(index)] = str(container_resources)
                    ws['O' + str(index)] = str(container_liveness_probe)
                    ws['P' + str(index)] = str(container_readiness_probe)
                    ws['Q' + str(index)] = str(container_volume_mounts)
                    index += 1
    wb.save('platform_info.xlsx')
